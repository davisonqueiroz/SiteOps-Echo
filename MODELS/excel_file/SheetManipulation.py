import pandas as pd
import os
import chardet
from openpyxl import load_workbook

class SheetManipulation:
    
    HEADERS = {
        "msp_offers" : ["Nome da IES", "ID da IES", "Nome do Campus","ID do Campus", "Nome do Curso" ],
        "exp_offers" : ["university_id", "campus_id", "name_from_university", "level", "kind"],
        "csv_kroton" : ["match_key", "campus_name", "max_payments", "full_price", "metadata_course_id"],
        "msp_campus" : ["ID da IES", "IES", "Nome de Exibição", "Nome", "Logradouro", "Número"],
        "exp_campus" : ["id", "name", "education_group_id", "metadata_code", "match_codes"]
    }
    def __init__(self,path,sheet_name = None):
        self.path = path
        self.sheet_name = sheet_name
        self.file_type = self.set_file_type()
        self.encoding = None
        self.delimiter = None
        self.sheet_type = None
        self.dtype = None
    
     #//////// dataframe configurations ////////

    def set_file_type(self):
        if self.path is None:
            raise FileNotFoundError(f"Caminho de arquivo inválido: {self.path}")
        extension = self._detect_extension(self.path)
        if extension not in ['.xlsx','.csv']:
            raise ValueError(f"Tipo de arquivo {extension} não suportado. Selecione um arquivo '.xlsx' ou '.csv' e tente novamente.")
        return extension
     
    def _detect_extension(self,path):
        return os.path.splitext(path)[1].lower()

    def load(self):
        if self._file_type_is_ready():
            if self.sheet_name is None and self.file_type == ".xlsx":
                self._set_sheet_name() 
            self._set_sheet_type()
            if self.file_type == ".xlsx":
                return self._load_xlsx()
            else:
                return self._load_csv()

    def _file_type_is_ready(self):
        if self.file_type is not None:
            return True
        else:
            return False
        
    def _set_sheet_type(self):
        headers_found = self._get_headers()
        for sheet_type,expected_headers in self.HEADERS.items():
            if all(header in headers_found for header in expected_headers):
                self.sheet_type = sheet_type
                return
        self.sheet_type =  "others"

    def _get_headers(self):
        if self.file_type == ".xlsx":
            workbook = load_workbook(self.path,read_only=True,data_only=True)
            sheet = workbook[self.sheet_name] if self.sheet_name else workbook.active()
            headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
            workbook.close()
            return headers
        else:
            with open(self.path, "r", encoding=self.encoding) as f:
                line = f.readline()
                delimiter = self.delimiter
                return line.strip().split(delimiter)
            
            #//////// xlsx configurations ////////

    def _load_xlsx(self):
        if not self.xlsx_is_ready():
            raise ValueError ("Informações de carregamento incorretas. Confira o dtype e sheet_name")
        if self.sheet_type == "msp_offers":
            self.set_msp_offers_dtype()
            try:
                dataframe = pd.read_excel(self.path,sheet_name=self.sheet_name,dtype=self.dtype)
                return self.adjust_percentages_in_msp(dataframe)
            except Exception as e:
                raise ValueError (f"Erro ao carregar planilha Excel: {e}")
        elif self.sheet_type == "exp_offers":
            try:
                return pd.read_excel(self.path,sheet_name=self.sheet_name)
            except Exception as e:
                raise ValueError (f"Erro ao carregar planilha Excel: {e}")
        elif self.sheet_type == "exp_campus":
            try:
                return pd.read_excel(self.path,sheet_name=self.sheet_name,dtype= self.dtype)
            except Exception as e:
                raise ValueError (f"Erro ao carregar planilha Excel: {e}")
        elif self.sheet_type == "others":
            try:
                return pd.read_excel(self.path,sheet_name=self.sheet_name)
            except Exception as e:
                raise ValueError (f"Erro ao carregar planilha Excel: {e}")

        
    def adjust_percentages_in_msp(self,dataframe):
        percentage_columns = ['Porcentagem de desconto da bolsa (Fixo/1 º Semestre)',
                             'Porcentagem total de desconto da bolsa\n(2º Semestre)',
                             'Porcentagem total de desconto da bolsa\n(3º Semestre)',
                             'Porcentagem total de desconto da bolsa\n(4º Semestre)',
                             'Porcentagem total de desconto da bolsa\n(5º Semestre)',
                             'Porcentagem total de desconto da bolsa\n(6º Semestre)',
                             'Porcentagem total de desconto da bolsa\n(7º Semestre)',
                             'Porcentagem total de desconto da bolsa\n(8º Semestre)',
                             'Porcentagem total de desconto da bolsa\n(9º Semestre)',
                             'Porcentagem total de desconto da bolsa\n(10º Semestre)',
                             'Porcentagem de desconto IES',
                             'Porcentagem de desconto IES (1º semestre)',
                             'Porcentagem de desconto IES (2º semestre)',
                             'Porcentagem de desconto IES (3º semestre)',
                             'Porcentagem de desconto IES (4º semestre)',
                             'Porcentagem de desconto IES (5º semestre)',
                             'Porcentagem de desconto IES (6º semestre)',
                             'Porcentagem de desconto IES (7º semestre)',
                             'Porcentagem de desconto IES (8º semestre)',
                             'Porcentagem de desconto IES (9º semestre)',
                             'Porcentagem de desconto IES (10º semestre)'
                             ]
        lambda_apply = lambda x: f"{x:.2f}"
        dataframe.update({
            col: dataframe[col].apply(lambda_apply)
            for col in percentage_columns
            if col in dataframe.columns
        })
        return dataframe
    
    def set_msp_offers_dtype(self):
        self.dtype = {
            'ID_POLO' : str,
            'COD_CURSO' : str,
            'Semestre de Ingresso': str,
        }

    def set_exp_campus_dtype(self):
        self.dtype = {
            'id' : str,
            'metadata_code' : str,
            'university_id' : str
        }

    def xlsx_is_ready(self):
        if self.sheet_name:
            return True
        elif self.sheet_name is None and self.sheet_type == "others":
            self._set_sheet_name()
            return self.sheet_name is not None
        else:
            return False
        
    def _set_sheet_name(self):
        if self.sheet_name is None and self.file_type == ".xlsx":
            wb = load_workbook(self.path, read_only=True, data_only=True)
            self.sheet_name = wb.sheetnames[0] 
            wb.close()


        #//////// csv configurations ////////

    def _load_csv(self):
        self.set_encoding()
        self.set_delimiter()
        if self.csv_is_ready():
            try:
                return pd.read_csv(self.path,self.delimiter,self.encoding)
            except Exception as e:
                raise ValueError (f"Erro ao carregar planilha Excel: {e}")

    def csv_is_ready(self):
        if self.delimiter and self.encoding is not None:
            return True
        else:
            return False

    def set_encoding(self):
        with open(self.path, "rb") as f:
            encod = chardet.detect(f.read(1000))
            self.encoding =  encod['encoding']
        
    def set_delimiter(self):
        if self.encoding is None:
            raise ValueError()
        with open(self.path, "r", encoding=self.encoding) as f:
            first_row = f.readline()
            if "\t" in first_row:
                self.delimiter = "\t"
            elif ";" in first_row:
                self.delimiter = ";"
            elif "," in first_row:
                self.delimiter = ","
    
    def convert_to_msp_offers_and_load(self, relation_headers : dict):

        msp_headers = ['Nome da IES', 'ID da IES', 'Nome do Campus',
       'ID do Campus', 'Nome do Curso', 'Grau', 'Modalidade',
       'Turno', 'Tipo de duração do curso', 'Duração do Curso',
       'Quantidade de Parcelas', 'Qual valor usar?\n% ou R$',
       'Mensalidade sem desconto', 'Mensalidade com desconto',
       'Porcentagem de desconto da bolsa (Fixo/1 º Semestre)',
       'Porcentagem total de desconto da bolsa\n(2º Semestre)',
       'Mensalidade com desconto\n(2º Semestre)', 'Mensalidade balcão',
       'Porcentagem de desconto IES', 'Data de Início da Oferta',
       'Data de Fim da Oferta', 'LIMITADA?', 'Quantidade de Vagas',
       'Semestre de Ingresso', 'Benefício 1 (Chave OSC)',
       'Benefício 2 (Chave OSC)', 'Avisos', 'Benefícios Extras', 'Campanha',
       'Frequência das aulas', 'Taxa de matrícula', 'Data de início das aulas',
       'Carga horária do Curso (em horas)', 'TCC Obrigatório?', 'Restrita?',
       'Tipo de restrição (systems:)', 'ecode_pool_name', 'COD CURSO',
       'COD IES', 'COD CAMPUS', 'COD TIPO GRAD', 'COD TURNO', 'COD CURSO PAI',
       'COD CAMPUS PAI', 'CONCURSO', 'CodCursoVest', 'CodCursoIES',
       'NomeCurso', 'TurnoMetadata', 'CURRICULO', 'CodCampus', 'CodCampanha',
       'affiliate_link', 'tags']
        
        self.set_msp_offers_dtype()

        dataframe_origin = pd.read_excel(self.path)
        dataframe_msp = pd.DataFrame(columns=msp_headers)
        for header,path_header in relation_headers.items():
            if path_header in dataframe_origin.columns:
                dataframe_msp[header] = dataframe_origin[path_header]
            else:
                raise KeyError (f"Coluna {path_header} inválida. Não encontrada na planilha.")
        dataframe_msp  = dataframe_msp.astype({
            col: tipo
            for col, tipo in self.dtype.items()
            if col in dataframe_msp.columns
        })
        return self.adjust_percentages_in_msp(dataframe_msp)
    
